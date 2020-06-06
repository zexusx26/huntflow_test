from openpyxl import load_workbook
from typing import List, Iterator, Any, Dict


class XLSXReaderException(Exception):
    pass


class XLSXReader:
    """ Класс чтеца файлов XLSX. """

    def __init__(self, fields: List['Field'], header_exists: bool = True):
        """ Инициализация чтеца XLSX файлов.

        :param fields: поля, XLSX файла
        :type fields: List[Field]
        :param header_exists: в файле есть заголовок, он будет игнорироваться, defaults to True
        :type header_exists: bool, optional
        """
        self.fields = fields
        self.header_exists = header_exists
        self.workbook = None
        self.sheet = None

    def __iter__(self) -> Iterator[Dict[str, Any]]:
        """ Итерироваться по строкам первого листа файла.

        :yield: итератор данных, извлеченных из строк
        :rtype: Iterator[Dict[str, Any]]
        """
        return self.read()

    def __enter__(self):
        """ При входе в менеджер контекста проверяет, что файл открыт.

        :raises XLSXReaderException: если файл не открыт
        """
        if self.workbook is None:
            raise XLSXReaderException('Workbook is not open.')
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        """ При выходе из контекстного менеджера закрывает файл. """
        self.close()

    def open(self, fp: str, sheet_number: int = 0) -> 'XLSXReader':
        """ Открыть файл для чтения.

        Если уже открыт файл для чтения, закрывает текущий, открывает новый.

        :param fp: путь до файла
        :type fp: str
        :param sheet_number: номер страницы, defaults to 1
        :type sheet_number: int, optional
        """
        if self.workbook is not None:
            self.close()
        self.workbook = load_workbook(filename=fp)
        self.sheet = self.workbook.worksheets[sheet_number]
        return self

    def close(self):
        """ Закрыть файл для чтения.

        :raises XLSXReaderException: если файл не открыт
        """
        if self.workbook is None:
            raise XLSXReaderException('Workbook is not open.')
        self.sheet = None
        self.workbook.close()
        self.workbook = None

    def read_row(self, row_number: int = 1) -> Dict[str, Any]:
        """ Прочитать строку.

        :param row_number: номер строки, defaults to 1
        :type row_number: int, optional
        :raises XLSXReaderException: если не открыт файл для чтения
        :return: данные, извлеченные из строки
        :rtype: Dict[str, Any]
        """
        if self.workbook is None:
            raise XLSXReaderException('Workbook is not open.')
        if self.header_exists:
            row_number = row_number + 1
        row = self.sheet[row_number]
        data = {}
        for col_number, field in enumerate(self.fields):
            col = row[col_number]
            value = field.parse(col.value)
            if value:
                data.update(value)
        return data

    # def read_sheet(self, sheet_number: int = 0, start_row: int = 1) -> Iterator[Dict[str, Any]]:
    #     """ Читать лист файла.

    #     Нумерация страниц начинается с 0, нумерация строк начинается с 1.

    #     :param start_row: строка, с которой начать чтение, defaults to 1
    #     :type start_row: int, optional
    #     :param sheet_number: лист, с которого начать чтение, defaults to 0
    #     :type sheet_number: int, optional
    #     :yield: итератор данных, извлеченных из строк
    #     :rtype: Iterator[Dict[str, Any]]
    #     """

    #     with self()

    #     workbook = load_workbook(filename=self.fp)
    #     sheet = workbook.worksheets[sheet_number]

    #     if self.header_exists:
    #         row_number = start_row + 1
    #     else:
    #         row_number = start_row

    #     while True:
    #         row = sheet[row_number]
    #         data = {}
    #         for col_number, field in enumerate(self.fields):
    #             col = row[col_number]
    #             value = field.parse(col.value)
    #             if value:
    #                 data.update(value)
    #         if data:
    #             yield data
    #         else:
    #             break
    #         row_number += 1
